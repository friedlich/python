import openpyxl
#读取的代码：
wb = openpyxl.load_workbook(r'c:/Users/asus/Desktop/Python/风变Python爬虫精进/爬虫阶段练习/6.19/Marvel1.xlsx')
# 第14行代码：调用openpyxl.load_workbook()函数，打开“Marvel.xlsx”文件。
sheet = wb['new title']
# 第15行代码：获取“Marvel.xlsx”工作薄中名为“new title”的工作表。
sheetname = wb.sheetnames
print(sheetname)
# 第16、17行代码：sheetnames是用来获取工作薄所有工作表的名字的。如果你不知道工作薄到底有几个工作表，就可以把工作表的名字都打印出来。
A1_cell = sheet['A1']
A1_value = A1_cell.value
# 第18-20行代码：把“new title”工作表中A1单元格赋值给A1_cell，再利用单元格value属性，就能打印出A1单元格的值。
print(A1_value)
